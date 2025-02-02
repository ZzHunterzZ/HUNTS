import json
import os
import subprocess
import openpyxl

def processar_sor(caminho_sor):
    """
    Processa o arquivo .sor usando o pyotdr e gera um JSON correspondente.
    """
    try:
        comando = ["pyotdr", caminho_sor]
        subprocess.run(comando, check=True)
        print(f"Processamento do arquivo {caminho_sor} concluído com sucesso.")
    except subprocess.CalledProcessError as e:
        print(f"Erro ao executar pyotdr: {e}")
        return False
    return True

def extrair_dados_json(nome_base):
    """
    Extrai informações do JSON gerado pelo pyotdr.
    """
    caminho_json = os.path.join(os.getcwd(), nome_base + "-dump.json")
    
    try:
        with open(caminho_json, "r") as arquivo_json:
            dados = json.load(arquivo_json)

        info = {
            "date/time": dados.get("FxdParams", {}).get("date/time", ""),
            "pulse width": dados.get("FxdParams", {}).get("pulse width", ""),
            "wavelength": dados.get("FxdParams", {}).get("wavelength", ""),
            "cable ID": dados.get("GenParams", {}).get("cable ID", ""),
            "location A": dados.get("GenParams", {}).get("location A", ""),
            "location B": dados.get("GenParams", {}).get("location B", ""),
            "operator": dados.get("GenParams", {}).get("operator", ""),
            "total loss": dados.get("KeyEvents", {}).get("Summary", {}).get("total loss", ""),
            "loss end": dados.get("KeyEvents", {}).get("Summary", {}).get("loss end", ""),
            "filename": nome_base + ".sor"
        }
        return info
    except FileNotFoundError:
        print(f"Arquivo JSON {caminho_json} não encontrado.")
    except json.JSONDecodeError:
        print(f"Erro ao decodificar o arquivo JSON {caminho_json}.")
    return {}

def atualizar_planilha(caminho_planilha, dados, sentido, primeiro_arquivo):
    """
    Atualiza a planilha Excel com os dados extraídos.
    """
    wb = openpyxl.load_workbook(caminho_planilha)
    ws = wb["OTDR"]

    if primeiro_arquivo:
        ws["G2"] = dados["date/time"]
        ws["G15"] = dados["pulse width"]
        ws["E14"] = dados["wavelength"]
        ws["F6"] = dados["cable ID"]
        ws["D6"] = dados["location A"]
        ws["D7"] = dados["location B"]
        ws["H3"] = dados["operator"]
    
    try:
        numero_fibra = int(dados["filename"].split("_")[1])
    except (IndexError, ValueError):
        print(f"Erro ao extrair número da fibra do arquivo {dados['filename']}")
        return
    
    linha = 19 + numero_fibra
    if sentido == "A-B":
        ws[f"D{linha}"] = dados["loss end"]
        ws[f"E{linha}"] = dados["total loss"]
    elif sentido == "B-A":
        ws[f"L{linha}"] = dados["loss end"]
        ws[f"M{linha}"] = dados["total loss"]
    
    for i in range(20, 308):
        if ws[f"D{i}"].value is None and ws[f"E{i}"].value is None:
            ws[f"H{i}"] = "F.O Ocupada/ Em serviço"
    
    wb.save(caminho_planilha)

def processar_arquivos_sor(pasta_sor, caminho_planilha, sentido):
    """
    Processa todos os arquivos .sor na pasta e atualiza a planilha.
    """
    arquivos_sor = sorted([f for f in os.listdir(pasta_sor) if f.endswith(".sor")])
    primeiro_arquivo = True
    
    for arquivo in arquivos_sor:
        caminho_arquivo = os.path.join(pasta_sor, arquivo)
        nome_base = os.path.splitext(arquivo)[0]
        
        if not processar_sor(caminho_arquivo):
            continue
        
        dados_extraidos = extrair_dados_json(nome_base)
        
        if dados_extraidos:
            atualizar_planilha(caminho_planilha, dados_extraidos, sentido, primeiro_arquivo)
            primeiro_arquivo = False
        
        # Remover arquivos temporários
        try:
            os.remove(os.path.join(os.getcwd(), nome_base + "-dump.json"))
            os.remove(os.path.join(os.getcwd(), nome_base + "-trace.dat"))
        except FileNotFoundError:
            pass

if __name__ == "__main__":
    pasta_sor = "data/arquivos_sor"
    caminho_planilha = "data/planilha.xlsx"
    sentido = "A-B"
    processar_arquivos_sor(pasta_sor, caminho_planilha, sentido)
